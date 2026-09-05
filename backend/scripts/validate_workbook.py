"""
CLI Tool: Validate Clinic Workbook
Checks that clinic_data.xlsx has all 12 required sheets and correct column headers.
"""
import sys
from pathlib import Path
import openpyxl

# Add backend directory to sys.path
backend_dir = Path(__file__).resolve().parent.parent
if str(backend_dir) not in sys.path:
    sys.path.insert(0, str(backend_dir))

from app.core.config import settings
from app.repositories.excel_schema import ALL_SHEETS, SHEET_COLUMNS

def validate_workbook(path: Path = settings.WORKBOOK_PATH) -> bool:
    print(f"[*] Validating clinic workbook at: {path}")
    if not path.exists():
        print(f"[!] ERROR: Workbook not found at {path}")
        return False

    wb = openpyxl.load_workbook(path, data_only=True)
    existing_sheets = wb.sheetnames
    is_valid = True

    print("\n--- Sheet & Header Validation ---")
    for sheet_name in ALL_SHEETS:
        if sheet_name not in existing_sheets:
            print(f"[x] MISSING SHEET: {sheet_name}")
            is_valid = False
            continue

        ws = wb[sheet_name]
        expected_cols = SHEET_COLUMNS.get(sheet_name, [])
        actual_cols = [cell.value for cell in ws[1] if cell.value is not None]

        missing_cols = [c for c in expected_cols if c not in actual_cols]
        if missing_cols:
            print(f"[!] {sheet_name}: Missing expected columns {missing_cols}")
            is_valid = False
        else:
            row_count = ws.max_row - 1 if ws.max_row > 1 else 0
            print(f"[✓] {sheet_name:20s} | {len(actual_cols):2d} cols | {row_count:4d} rows")

    if is_valid:
        print("\n[✓] SUCCESS: All sheets and headers match schema specifications.")
    else:
        print("\n[!] WARNING: Schema mismatches found.")
    return is_valid

if __name__ == "__main__":
    success = validate_workbook()
    sys.exit(0 if success else 1)
