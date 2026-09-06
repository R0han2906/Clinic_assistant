import os
import json
import shutil
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.locking import get_workbook_lock
from app.core.exceptions import (
    WorkbookWriteError, SlotConflictError
)
from app.models.patient import PatientResponse, PatientCreate
from app.models.visit import VisitResponse, VisitCreate
from app.models.dentist import DentistResponse, DentistCreate
from app.models.availability import WorkingScheduleItem, LeaveItem, LeaveCreate, ScheduleUpdate
from app.models.appointment import (
    AppointmentResponse, AppointmentCreate, AppointmentStatus,
    PaymentReminderResponse
)
from app.models.audit import AuditLogEntry
from app.models.treatment import Treatment, TreatmentCreate, TreatmentUpdate, TreatmentResponse
from app.models.staff import StaffMember, StaffCreate, StaffUpdate
from app.models.sales import SaleResponse, SaleCreate, SaleSummary, PaymentMethodResponse
from app.models.purchase import PurchaseResponse, PurchaseCreate, VendorResponse, VendorCreate
from app.models.inventory import InventoryResponse, InventoryCreate, InventoryUpdate
from app.models.peripheral import PeripheralResponse, PeripheralCreate, PeripheralUpdate
from app.models.medical_checkup import (
    MedicalCheckupCreate, MedicalCheckupResponse, ToothFinding
)
from app.models.patient_request import (
    PatientRequestCreate, PatientRequestResponse, PatientRequestStatus
)
from app.repositories.base import BaseClinicRepository
from app.repositories.excel_schema import (
    ALL_SHEETS, SHEET_COLUMNS,
    SHEET_PATIENTS, SHEET_VISITS, SHEET_DENTISTS, SHEET_AVAILABILITY,
    SHEET_LEAVES, SHEET_APPOINTMENTS, SHEET_STAFF, SHEET_AUDIT, SHEET_METADATA,
    SHEET_CHECKUPS, SHEET_TREATMENTS, SHEET_PATIENT_REQUESTS,
    DEFAULT_DENTISTS, DEFAULT_AVAILABILITY, DEFAULT_METADATA,
    DEFAULT_TREATMENTS, DEFAULT_PATIENTS, DEFAULT_APPOINTMENTS
)


class ExcelClinicRepository(BaseClinicRepository):
    """
    OpenPyXL-based repository implementation for clinic pilot storage.
    Enforces atomic file writes, automatic backups, and thread/process locking.

    Architecture: Every public method acquires the workbook lock ONCE, reads
    all sheet data into memory, performs all operations on the in-memory dict
    via private _unlocked helpers, then writes back — releasing the lock once.
    This eliminates all re-entrant lock deadlock scenarios.
    """

    def __init__(
        self,
        workbook_path: Optional[Path] = None,
        backup_dir: Optional[Path] = None,
        auto_backup: Optional[bool] = None,
    ):
        self.workbook_path = Path(workbook_path or settings.WORKBOOK_PATH).resolve()
        self.backup_dir = Path(backup_dir or settings.BACKUP_DIR).resolve()
        self.lock_path = self.workbook_path.parent / f"{self.workbook_path.name}.lock"
        self.auto_backup = (
            auto_backup
            if auto_backup is not None
            else getattr(settings, "AUTO_BACKUP_ON_SAVE", False)
        )
        self._treatments = {}
        self._staff = {}
        self._sales = {}
        self._purchases = {}
        self._inventory = {}
        self._vendors = {}
        self._payment_methods = {}
        self._peripherals = {}
        self.initialize_storage()
        self._ensure_runtime_defaults()

    # =========================================================================
    # INITIALIZATION & HEALTH
    # =========================================================================

    def purge_redundant_backups(self) -> int:
        """Removes orphaned auto-generated backup files from the backup directory."""
        removed = 0
        if self.backup_dir.exists():
            for f in self.backup_dir.glob("clinic_data_20*.xlsx"):
                try:
                    f.unlink()
                    removed += 1
                except Exception:
                    pass
            for f in self.backup_dir.glob("pre_reset_backup_*.xlsx"):
                try:
                    f.unlink()
                    removed += 1
                except Exception:
                    pass
        return removed

    def initialize_storage(self) -> None:
        """Creates and formats the Excel workbook if it does not already exist."""
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        self.purge_redundant_backups()
        with get_workbook_lock(self.lock_path):
            if not self.workbook_path.exists():
                self._create_fresh_workbook()

    def check_health(self) -> Dict[str, Any]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            backups = list(self.backup_dir.glob("*.xlsx"))
            return {
                "status": "healthy",
                "workbook_exists": self.workbook_path.exists(),
                "workbook_path": str(self.workbook_path),
                "total_patients": len(all_data.get(SHEET_PATIENTS, [])),
                "total_appointments": len(all_data.get(SHEET_APPOINTMENTS, [])),
                "total_dentists": len(all_data.get(SHEET_DENTISTS, [])),
                "total_backups": len(backups),
                "last_checked": datetime.now().isoformat()
            }

    # =========================================================================
    # WORKBOOK I/O INTERNALS
    # =========================================================================

    def _create_fresh_workbook(self) -> None:
        """Called only when workbook does not exist, already under lock."""
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")

        for sheet_name in ALL_SHEETS:
            ws = wb.create_sheet(title=sheet_name)
            columns = SHEET_COLUMNS.get(sheet_name, [])
            ws.append(columns)
            for col_num, _ in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Seed initial dentists
        ws_dentists = wb[SHEET_DENTISTS]
        for d in DEFAULT_DENTISTS:
            ws_dentists.append([
                d["dentist_id"], d["name"], d["specialty"], d["phone"],
                d["email"], d["color_code"], "TRUE" if d["is_active"] else "FALSE",
                d["created_at"]
            ])

        # Seed initial availability
        ws_avail = wb[SHEET_AVAILABILITY]
        for a in DEFAULT_AVAILABILITY:
            ws_avail.append([
                a["availability_id"], a["dentist_id"], a["day_of_week"],
                a["start_time"], a["end_time"], a["break_start"], a["break_end"],
                "TRUE" if a["is_working_day"] else "FALSE"
            ])

        # Seed initial metadata
        ws_meta = wb[SHEET_METADATA]
        for m in DEFAULT_METADATA:
            ws_meta.append([m["key"], m["value"], m["updated_at"]])

        self._auto_fit_columns(wb)
        self._save_workbook_atomic(wb)

    def seed_demo_data(self) -> None:
        """
        Explicitly populates demo patients and appointments for local frontend testing.
        Never called automatically during test fixture creation.
        """
        with get_workbook_lock(self.lock_path):
            wb = openpyxl.load_workbook(self.workbook_path)
            try:
                # Seed initial patients if sheet is empty
                ws_patients = wb[SHEET_PATIENTS]
                if ws_patients.max_row <= 1:
                    for p in DEFAULT_PATIENTS:
                        ws_patients.append([
                            p["patient_id"], p["full_name"], p["dob_or_age"], p["phone"],
                            p["email"], p["emergency_contact"], p["gender"], p["address"],
                            p["allergies"], p["medical_conditions"], p["consent_status"],
                            p["created_at"], p["updated_at"]
                        ])

                # Seed initial appointments if sheet is empty
                ws_appts = wb[SHEET_APPOINTMENTS]
                if ws_appts.max_row <= 1:
                    for ap in DEFAULT_APPOINTMENTS:
                        ws_appts.append([
                            ap["appointment_id"], ap["patient_id"], ap["dentist_id"],
                            ap["date"], ap["start_time"], ap["end_time"], ap["booking_time"],
                            ap["treatment_name"], ap["source"], ap["payment_status"],
                            ap["bill_number"], ap["clinical_notes"], ap["status"],
                            ap["reason"], ap["notes"], ap["created_at"], ap["updated_at"]
                        ])

                self._auto_fit_columns(wb)
                self._save_workbook_atomic(wb)
            finally:
                wb.close()

    def _auto_fit_columns(self, wb: openpyxl.Workbook) -> None:
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    def _save_workbook_atomic(self, wb: openpyxl.Workbook) -> None:
        """
        Saves the workbook to a temp file, atomically replaces the original.
        Only creates a timestamped backup if self.auto_backup is explicitly True.
        Always called from within an existing lock — no lock acquired here.
        """
        temp_path = self.workbook_path.with_name(
            f"{self.workbook_path.stem}.tmp{self.workbook_path.suffix}"
        )
        try:
            wb.save(temp_path)
            wb.close()

            if self.auto_backup and self.workbook_path.exists():
                self.backup_dir.mkdir(parents=True, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                backup_file = self.backup_dir / f"clinic_data_{timestamp}.xlsx"
                shutil.copy2(self.workbook_path, backup_file)

            # Atomic replace — Windows-compatible
            os.replace(str(temp_path), str(self.workbook_path))
        except Exception as e:
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass
            raise WorkbookWriteError(f"Failed to atomically save workbook: {str(e)}")

    def create_manual_backup(self) -> Path:
        """Creates an explicit, intentional timestamped backup of clinic_data.xlsx in backup_dir."""
        with get_workbook_lock(self.lock_path):
            if not self.workbook_path.exists():
                raise WorkbookWriteError("Cannot create backup: clinic_data.xlsx does not exist.")
            self.backup_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = self.backup_dir / f"clinic_data_backup_{timestamp}.xlsx"
            shutil.copy2(self.workbook_path, backup_file)
            return backup_file

    def _read_all_sheets(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Reads all rows from all sheets into a dict of lists.
        Must be called inside an existing lock.
        """
        wb = openpyxl.load_workbook(self.workbook_path, data_only=True)
        try:
            data: Dict[str, List[Dict[str, Any]]] = {}
            for sheet_name in ALL_SHEETS:
                if sheet_name not in wb.sheetnames:
                    data[sheet_name] = []
                    continue
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    data[sheet_name] = []
                    continue
                headers = [
                    str(h).strip() if h is not None else f"col_{i}"
                    for i, h in enumerate(rows[0])
                ]
                sheet_rows = []
                for row in rows[1:]:
                    if all(c is None for c in row):
                        continue
                    row_dict: Dict[str, str] = {}
                    for i, header in enumerate(headers):
                        val = row[i] if i < len(row) else None
                        row_dict[header] = "" if val is None else str(val).strip()
                    sheet_rows.append(row_dict)
                data[sheet_name] = sheet_rows
            return data
        finally:
            wb.close()

    def _write_all_sheets(self, all_data: Dict[str, List[Dict[str, Any]]]) -> None:
        """
        Re-generates and atomically saves the workbook from the given data dict.
        Must be called inside an existing lock.
        """
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")

        for sheet_name in ALL_SHEETS:
            ws = wb.create_sheet(title=sheet_name)
            columns = SHEET_COLUMNS.get(sheet_name, [])
            ws.append(columns)
            for col_num in range(1, len(columns) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row_dict in all_data.get(sheet_name, []):
                row_vals = [row_dict.get(col, "") for col in columns]
                ws.append(row_vals)

        self._auto_fit_columns(wb)
        self._save_workbook_atomic(wb)

    # =========================================================================
    # SEQUENCE ID HELPER
    # =========================================================================

    def _next_sequence(self, rows: List[Dict[str, Any]], id_field: str) -> int:
        """
        Returns max sequence number + 1 by parsing existing ID values.
        Collision-safe even after deletions — never uses len(rows).
        """
        nums = []
        for r in rows:
            raw = r.get(id_field, "")
            parts = raw.split("-")
            if parts and parts[-1].isdigit():
                nums.append(int(parts[-1]))
        return max(nums, default=0) + 1

    # =========================================================================
    # PATIENTS — PUBLIC (lock) + PRIVATE (no lock)
    # =========================================================================

    def list_patients(self) -> List[PatientResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            return self._list_patients_unlocked(all_data)

    def _list_patients_unlocked(self, all_data: Dict) -> List[PatientResponse]:
        """Builds PatientResponse list from already-loaded sheet data. No lock."""
        results = []
        for r in all_data.get(SHEET_PATIENTS, []):
            results.append(PatientResponse(
                patient_id=r.get("patient_id", ""),
                full_name=r.get("full_name", ""),
                dob_or_age=r.get("dob_or_age", ""),
                phone=r.get("phone", ""),
                email=r.get("email") or None,
                emergency_contact=r.get("emergency_contact") or None,
                gender=r.get("gender") or None,
                address=r.get("address") or None,
                allergies=r.get("allergies") or None,
                medical_conditions=r.get("medical_conditions") or None,
                consent_status=r.get("consent_status", "acknowledged"),
                created_at=r.get("created_at", ""),
                updated_at=r.get("updated_at", "")
            ))
        return results

    def get_patient(self, patient_id: str) -> Optional[PatientResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            return self._get_patient_unlocked(all_data, patient_id)

    def _get_patient_unlocked(
        self, all_data: Dict, patient_id: str
    ) -> Optional[PatientResponse]:
        """Finds a patient by ID from already-loaded data. No lock."""
        for p in self._list_patients_unlocked(all_data):
            if p.patient_id.lower() == patient_id.lower():
                return p
        return None

    def find_patients(self, query: str) -> List[PatientResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            patients = self._list_patients_unlocked(all_data)
        query_clean = query.strip().lower()
        if not query_clean:
            return patients
        return [
            p for p in patients
            if query_clean in p.patient_id.lower()
            or query_clean in p.full_name.lower()
            or query_clean in p.phone.replace(" ", "").replace("-", "")
            or (p.email and query_clean in p.email.lower())
        ]

    def create_patient(self, patient_data: PatientCreate) -> PatientResponse:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            patients_rows = all_data.get(SHEET_PATIENTS, [])

            seq = self._next_sequence(patients_rows, "patient_id")
            patient_id = f"PAT-{seq:06d}"
            now_iso = datetime.now().isoformat()

            new_row = {
                "patient_id": patient_id,
                "full_name": patient_data.full_name,
                "dob_or_age": patient_data.dob_or_age,
                "phone": patient_data.phone,
                "email": patient_data.email or "",
                "emergency_contact": patient_data.emergency_contact or "",
                "gender": patient_data.gender or "",
                "address": patient_data.address or "",
                "allergies": patient_data.allergies or "",
                "medical_conditions": patient_data.medical_conditions or "",
                "consent_status": patient_data.consent_status,
                "created_at": now_iso,
                "updated_at": now_iso
            }
            patients_rows.append(new_row)
            all_data[SHEET_PATIENTS] = patients_rows

            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "staff_reception",
                "entity_type": "PATIENT",
                "entity_id": patient_id,
                "action": "CREATE",
                "details": f"Registered patient {patient_data.full_name} ({patient_data.phone})"
            })
            all_data[SHEET_AUDIT] = audit_rows

            self._write_all_sheets(all_data)

        return PatientResponse(
            patient_id=patient_id,
            full_name=patient_data.full_name,
            dob_or_age=patient_data.dob_or_age,
            phone=patient_data.phone,
            email=patient_data.email,
            emergency_contact=patient_data.emergency_contact,
            gender=patient_data.gender,
            address=patient_data.address,
            allergies=patient_data.allergies,
            medical_conditions=patient_data.medical_conditions,
            consent_status=patient_data.consent_status,
            created_at=now_iso,
            updated_at=now_iso
        )

    def update_patient(self, patient_id: str, updates: Dict[str, Any]) -> Optional[PatientResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            patients_rows = all_data.get(SHEET_PATIENTS, [])
            target = None
            for r in patients_rows:
                if r.get("patient_id", "").lower() == patient_id.lower():
                    target = r
                    break
            if not target:
                return None

            now_iso = datetime.now().isoformat()
            for k, v in updates.items():
                if k in SHEET_COLUMNS[SHEET_PATIENTS] and k not in ["patient_id", "created_at"]:
                    target[k] = "" if v is None else str(v)
            target["updated_at"] = now_iso

            self._write_all_sheets(all_data)

        return PatientResponse(
            patient_id=target["patient_id"],
            full_name=target["full_name"],
            dob_or_age=target["dob_or_age"],
            phone=target["phone"],
            email=target.get("email") or None,
            emergency_contact=target.get("emergency_contact") or None,
            gender=target.get("gender") or None,
            address=target.get("address") or None,
            allergies=target.get("allergies") or None,
            medical_conditions=target.get("medical_conditions") or None,
            consent_status=target.get("consent_status", "acknowledged"),
            created_at=target["created_at"],
            updated_at=target["updated_at"]
        )

    # =========================================================================
    # VISITS
    # =========================================================================

    def list_visits_for_patient(self, patient_id: str) -> List[VisitResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            visits_rows = all_data.get(SHEET_VISITS, [])
            dentists_map = {
                d.get("dentist_id"): d.get("name")
                for d in all_data.get(SHEET_DENTISTS, [])
            }

        results = []
        for r in visits_rows:
            if r.get("patient_id", "").lower() == patient_id.lower():
                results.append(VisitResponse(
                    visit_id=r.get("visit_id", ""),
                    patient_id=r.get("patient_id", ""),
                    visit_date=r.get("visit_date", ""),
                    dentist_id=r.get("dentist_id", ""),
                    dentist_name=dentists_map.get(r.get("dentist_id", "")),
                    visit_type=r.get("visit_type", ""),
                    summary=r.get("summary", ""),
                    follow_up_recommendation=r.get("follow_up_recommendation") or None,
                    created_at=r.get("created_at", "")
                ))
        return sorted(results, key=lambda v: v.visit_date, reverse=True)

    def create_visit(self, visit_data: VisitCreate) -> VisitResponse:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            visits_rows = all_data.get(SHEET_VISITS, [])
            dentists_map = {
                d.get("dentist_id"): d.get("name")
                for d in all_data.get(SHEET_DENTISTS, [])
            }

            seq = self._next_sequence(visits_rows, "visit_id")
            visit_id = f"VIS-{seq:06d}"
            now_iso = datetime.now().isoformat()

            new_row = {
                "visit_id": visit_id,
                "patient_id": visit_data.patient_id,
                "visit_date": visit_data.visit_date,
                "dentist_id": visit_data.dentist_id,
                "visit_type": visit_data.visit_type,
                "summary": visit_data.summary,
                "follow_up_recommendation": visit_data.follow_up_recommendation or "",
                "created_at": now_iso
            }
            visits_rows.append(new_row)
            all_data[SHEET_VISITS] = visits_rows

            self._write_all_sheets(all_data)

        return VisitResponse(
            visit_id=visit_id,
            patient_id=visit_data.patient_id,
            visit_date=visit_data.visit_date,
            dentist_id=visit_data.dentist_id,
            dentist_name=dentists_map.get(visit_data.dentist_id),
            visit_type=visit_data.visit_type,
            summary=visit_data.summary,
            follow_up_recommendation=visit_data.follow_up_recommendation,
            created_at=now_iso
        )

    # =========================================================================
    # DENTISTS — PUBLIC (lock) + PRIVATE (no lock)
    # =========================================================================

    def list_dentists(self, active_only: bool = True) -> List[DentistResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            return self._list_dentists_unlocked(all_data, active_only)

    def _list_dentists_unlocked(
        self, all_data: Dict, active_only: bool = True
    ) -> List[DentistResponse]:
        """Builds DentistResponse list from already-loaded data. No lock."""
        results = []
        for r in all_data.get(SHEET_DENTISTS, []):
            is_active = str(r.get("is_active", "TRUE")).upper() in ["TRUE", "1", "YES"]
            if active_only and not is_active:
                continue
            results.append(DentistResponse(
                dentist_id=r.get("dentist_id", ""),
                name=r.get("name", ""),
                specialty=r.get("specialty", ""),
                phone=r.get("phone") or None,
                email=r.get("email") or None,
                color_code=r.get("color_code", "#2B6CB0"),
                is_active=is_active,
                created_at=r.get("created_at", "")
            ))
        return results

    def get_dentist(self, dentist_id: str) -> Optional[DentistResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            return self._get_dentist_unlocked(all_data, dentist_id)

    def get_visit_by_appointment(self, appointment_id: str) -> Optional[VisitResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            visits_rows = all_data.get(SHEET_VISITS, [])
            dentists_map = {
                d.get("dentist_id"): d.get("name")
                for d in all_data.get(SHEET_DENTISTS, [])
            }
            for r in visits_rows:
                if str(r.get("appointment_id", "")).lower() == appointment_id.lower():
                    return VisitResponse(
                        visit_id=r.get("visit_id", ""),
                        patient_id=r.get("patient_id", ""),
                        visit_date=r.get("visit_date", ""),
                        dentist_id=r.get("dentist_id", ""),
                        dentist_name=dentists_map.get(r.get("dentist_id", "")),
                        visit_type=r.get("visit_type", ""),
                        summary=r.get("summary", ""),
                        follow_up_recommendation=r.get("follow_up_recommendation") or None,
                        created_at=r.get("created_at", "")
                    )
            apt_rows = all_data.get(SHEET_APPOINTMENTS, [])
            target_apt = next((a for a in apt_rows if a.get("appointment_id", "").lower() == appointment_id.lower()), None)
            if target_apt:
                p_id = target_apt.get("patient_id")
                a_date = target_apt.get("date")
                for r in reversed(visits_rows):
                    if r.get("patient_id") == p_id and r.get("visit_date") == a_date:
                        return VisitResponse(
                            visit_id=r.get("visit_id", ""),
                            patient_id=r.get("patient_id", ""),
                            visit_date=r.get("visit_date", ""),
                            dentist_id=r.get("dentist_id", ""),
                            dentist_name=dentists_map.get(r.get("dentist_id", "")),
                            visit_type=r.get("visit_type", ""),
                            summary=r.get("summary", ""),
                            follow_up_recommendation=r.get("follow_up_recommendation") or None,
                            created_at=r.get("created_at", "")
                        )
            return None

    def _get_dentist_unlocked(
        self, all_data: Dict, dentist_id: str
    ) -> Optional[DentistResponse]:
        """Finds a dentist by ID or short alias (d1, d2, d3). No lock."""
        id_lower = dentist_id.lower()
        alias_map = {
            "d1": "doc-000001",
            "d2": "doc-000002",
            "d3": "doc-000003",
        }
        target_id = alias_map.get(id_lower, id_lower)
        for d in self._list_dentists_unlocked(all_data, active_only=False):
            if d.dentist_id.lower() == target_id or d.dentist_id.lower() == id_lower:
                return d
        return None

    def create_dentist(self, dentist_data: DentistCreate) -> DentistResponse:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            dentists_rows = all_data.get(SHEET_DENTISTS, [])

            seq = self._next_sequence(dentists_rows, "dentist_id")
            dentist_id = f"DOC-{seq:06d}"
            now_iso = datetime.now().isoformat()

            new_row = {
                "dentist_id": dentist_id,
                "name": dentist_data.name,
                "specialty": dentist_data.specialty,
                "phone": dentist_data.phone or "",
                "email": dentist_data.email or "",
                "color_code": dentist_data.color_code,
                "is_active": "TRUE" if dentist_data.is_active else "FALSE",
                "created_at": now_iso
            }
            dentists_rows.append(new_row)
            all_data[SHEET_DENTISTS] = dentists_rows

            self._write_all_sheets(all_data)

        return DentistResponse(
            dentist_id=dentist_id,
            name=dentist_data.name,
            specialty=dentist_data.specialty,
            phone=dentist_data.phone,
            email=dentist_data.email,
            color_code=dentist_data.color_code,
            is_active=dentist_data.is_active,
            created_at=now_iso
        )

    # =========================================================================
    # AVAILABILITY / SCHEDULES
    # =========================================================================

    def list_schedules_for_dentist(self, dentist_id: str) -> List[WorkingScheduleItem]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            return self._list_schedules_unlocked(all_data, dentist_id)

    def _list_schedules_unlocked(
        self, all_data: Dict, dentist_id: str
    ) -> List[WorkingScheduleItem]:
        """Reads schedules for a dentist from already-loaded data. No lock."""
        results = []
        for r in all_data.get(SHEET_AVAILABILITY, []):
            if r.get("dentist_id", "").lower() == dentist_id.lower():
                is_working = str(r.get("is_working_day", "TRUE")).upper() in ["TRUE", "1", "YES"]
                results.append(WorkingScheduleItem(
                    availability_id=r.get("availability_id"),
                    dentist_id=r.get("dentist_id", ""),
                    day_of_week=int(r.get("day_of_week", 0)),
                    start_time=r.get("start_time", "09:00"),
                    end_time=r.get("end_time", "17:00"),
                    break_start=r.get("break_start") or None,
                    break_end=r.get("break_end") or None,
                    is_working_day=is_working
                ))
        return results

    def update_schedule_for_day(
        self,
        dentist_id: str,
        day_of_week: int,
        updates: ScheduleUpdate
    ) -> Optional[WorkingScheduleItem]:
        """
        Updates a dentist's working schedule for a specific day of the week.
        Creates a new row if one does not yet exist for that day.
        """
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            avail_rows = all_data.get(SHEET_AVAILABILITY, [])

            target = None
            for r in avail_rows:
                if (
                    r.get("dentist_id", "").lower() == dentist_id.lower()
                    and str(r.get("day_of_week", "")) == str(day_of_week)
                ):
                    target = r
                    break

            now_iso = datetime.now().isoformat()

            if target:
                # Update existing row
                target["start_time"] = updates.start_time
                target["end_time"] = updates.end_time
                target["break_start"] = updates.break_start or ""
                target["break_end"] = updates.break_end or ""
                target["is_working_day"] = "TRUE" if updates.is_working_day else "FALSE"
            else:
                # Create new schedule row for this day
                seq = self._next_sequence(avail_rows, "availability_id")
                target = {
                    "availability_id": f"AVL-{seq:06d}",
                    "dentist_id": dentist_id,
                    "day_of_week": str(day_of_week),
                    "start_time": updates.start_time,
                    "end_time": updates.end_time,
                    "break_start": updates.break_start or "",
                    "break_end": updates.break_end or "",
                    "is_working_day": "TRUE" if updates.is_working_day else "FALSE"
                }
                avail_rows.append(target)

            all_data[SHEET_AVAILABILITY] = avail_rows

            audit_rows = all_data.get(SHEET_AUDIT, [])
            day_names = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "staff_admin",
                "entity_type": "SCHEDULE",
                "entity_id": dentist_id,
                "action": "UPDATE",
                "details": (
                    f"Updated schedule for {dentist_id} on {day_names[day_of_week]}: "
                    f"{updates.start_time}-{updates.end_time}, "
                    f"working={updates.is_working_day}"
                )
            })
            all_data[SHEET_AUDIT] = audit_rows

            self._write_all_sheets(all_data)

        return WorkingScheduleItem(
            availability_id=target.get("availability_id"),
            dentist_id=dentist_id,
            day_of_week=day_of_week,
            start_time=updates.start_time,
            end_time=updates.end_time,
            break_start=updates.break_start or None,
            break_end=updates.break_end or None,
            is_working_day=updates.is_working_day
        )

    def list_leaves_for_dentist(self, dentist_id: str) -> List[LeaveItem]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            return self._list_leaves_unlocked(all_data, dentist_id)

    def _list_leaves_unlocked(self, all_data: Dict, dentist_id: str) -> List[LeaveItem]:
        """Reads leaves for a dentist from already-loaded data. No lock."""
        results = []
        for r in all_data.get(SHEET_LEAVES, []):
            if r.get("dentist_id", "").lower() == dentist_id.lower():
                results.append(LeaveItem(
                    leave_id=r.get("leave_id"),
                    dentist_id=r.get("dentist_id", ""),
                    start_date=r.get("start_date", ""),
                    end_date=r.get("end_date", ""),
                    reason=r.get("reason") or None,
                    created_at=r.get("created_at") or None
                ))
        return sorted(results, key=lambda l: l.start_date)

    def create_leave_for_dentist(self, dentist_id: str, leave_data: LeaveCreate) -> LeaveItem:
        """Blocks a dentist for a date range. Persisted in the Leaves sheet."""
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            leaves_rows = all_data.get(SHEET_LEAVES, [])

            seq = self._next_sequence(leaves_rows, "leave_id")
            leave_id = f"LVE-{seq:06d}"
            now_iso = datetime.now().isoformat()

            new_row = {
                "leave_id": leave_id,
                "dentist_id": dentist_id,
                "start_date": leave_data.start_date,
                "end_date": leave_data.end_date,
                "reason": leave_data.reason or "",
                "created_at": now_iso
            }
            leaves_rows.append(new_row)
            all_data[SHEET_LEAVES] = leaves_rows

            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "staff_admin",
                "entity_type": "LEAVE",
                "entity_id": dentist_id,
                "action": "CREATE",
                "details": (
                    f"Leave for {dentist_id} from {leave_data.start_date} "
                    f"to {leave_data.end_date}. Reason: {leave_data.reason or 'N/A'}"
                )
            })
            all_data[SHEET_AUDIT] = audit_rows

            self._write_all_sheets(all_data)

        return LeaveItem(
            leave_id=leave_id,
            dentist_id=dentist_id,
            start_date=leave_data.start_date,
            end_date=leave_data.end_date,
            reason=leave_data.reason,
            created_at=now_iso
        )

    # =========================================================================
    # APPOINTMENTS — PUBLIC (lock) + PRIVATE (no lock)
    # =========================================================================

    def list_appointments(
        self,
        date: Optional[str] = None,
        dentist_id: Optional[str] = None,
        patient_id: Optional[str] = None,
        status: Optional[AppointmentStatus] = None
    ) -> List[AppointmentResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            return self._list_appointments_unlocked(all_data, date, dentist_id, patient_id, status)

    def _list_appointments_unlocked(
        self,
        all_data: Dict,
        date: Optional[str] = None,
        dentist_id: Optional[str] = None,
        patient_id: Optional[str] = None,
        status: Optional[AppointmentStatus] = None
    ) -> List[AppointmentResponse]:
        """Builds AppointmentResponse list from already-loaded data. No lock."""
        apt_rows = all_data.get(SHEET_APPOINTMENTS, [])
        patients_map = {p.get("patient_id"): p for p in all_data.get(SHEET_PATIENTS, [])}
        dentists_map = {d.get("dentist_id"): d.get("name") for d in all_data.get(SHEET_DENTISTS, [])}

        results = []
        for r in apt_rows:
            if date and r.get("date") != date:
                continue
            if dentist_id and r.get("dentist_id", "").lower() != dentist_id.lower():
                continue
            if patient_id and r.get("patient_id", "").lower() != patient_id.lower():
                continue
            if status and r.get("status", "").lower() != status.value.lower():
                continue

            p_info = patients_map.get(r.get("patient_id", ""), {})
            raw_status = r.get("status", "confirmed").lower()
            try:
                apt_status = AppointmentStatus(raw_status)
            except ValueError:
                apt_status = AppointmentStatus.CONFIRMED

            results.append(AppointmentResponse(
                appointment_id=r.get("appointment_id", ""),
                patient_id=r.get("patient_id", ""),
                patient_name=p_info.get("full_name"),
                patient_phone=p_info.get("phone"),
                dentist_id=r.get("dentist_id", ""),
                dentist_name=dentists_map.get(r.get("dentist_id", "")),
                date=r.get("date", ""),
                start_time=r.get("start_time", ""),
                end_time=r.get("end_time", ""),
                treatment_name=r.get("treatment_name") or "General Checkup",
                source=r.get("source") or "MANUAL APPOINTMENT",
                payment_status=r.get("payment_status") or "UNPAID",
                bill_number=r.get("bill_number") or None,
                clinical_notes=r.get("clinical_notes") or None,
                status=apt_status,
                reason=r.get("reason") or None,
                notes=r.get("notes") or None,
                booking_time=r.get("booking_time") or r.get("created_at", ""),
                created_at=r.get("created_at", ""),
                updated_at=r.get("updated_at", "")
            ))
        return sorted(results, key=lambda a: (a.date, a.start_time))

    def get_appointment(self, appointment_id: str) -> Optional[AppointmentResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            return self._get_appointment_unlocked(all_data, appointment_id)

    def _get_appointment_unlocked(
        self, all_data: Dict, appointment_id: str
    ) -> Optional[AppointmentResponse]:
        """Finds a single appointment by ID from already-loaded data. No lock."""
        for apt in self._list_appointments_unlocked(all_data):
            if apt.appointment_id.lower() == appointment_id.lower():
                return apt
        return None

    def create_appointment(self, appointment_data: AppointmentCreate) -> AppointmentResponse:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            apt_rows = all_data.get(SHEET_APPOINTMENTS, [])
            patients_map = {p.get("patient_id"): p for p in all_data.get(SHEET_PATIENTS, [])}
            dentists_map = {d.get("dentist_id"): d.get("name") for d in all_data.get(SHEET_DENTISTS, [])}

            # Conflict double-check under lock using in-memory data (no extra lock)
            alias_map = {"d1": "DOC-000001", "d2": "DOC-000002", "d3": "DOC-000003"}
            target_dentist_id = alias_map.get(appointment_data.dentist_id.lower(), appointment_data.dentist_id)

            for r in apt_rows:
                r_doc = r.get("dentist_id", "").lower()
                if (
                    (r_doc == appointment_data.dentist_id.lower() or r_doc == target_dentist_id.lower())
                    and r.get("date") == appointment_data.date
                    and r.get("status", "").lower().replace("_", "-") in [
                        "confirmed", "pending", "registered", "scheduled", "checked-in", "in-progress"
                    ]
                ):
                    existing_start = r.get("start_time", "")
                    existing_end = r.get("end_time", "")
                    if not (
                        appointment_data.end_time <= existing_start
                        or appointment_data.start_time >= existing_end
                    ):
                        raise SlotConflictError(
                            f"Dentist is already booked on {appointment_data.date} "
                            f"from {existing_start} to {existing_end}."
                        )

            seq = self._next_sequence(apt_rows, "appointment_id")
            apt_id = f"APT-{seq:06d}"
            now_iso = datetime.now().isoformat()
            booking_timestamp = appointment_data.booking_time or now_iso
            treatment = appointment_data.treatment_name or "General Checkup"
            bill_ref = appointment_data.bill_number or f"Bill #{10100 + seq}"
            stat_val = appointment_data.status or "confirmed"
            if hasattr(stat_val, "value"):
                stat_val = stat_val.value

            new_row = {
                "appointment_id": apt_id,
                "patient_id": appointment_data.patient_id,
                "dentist_id": target_dentist_id,
                "date": appointment_data.date,
                "start_time": appointment_data.start_time,
                "end_time": appointment_data.end_time,
                "booking_time": booking_timestamp,
                "treatment_name": treatment,
                "source": appointment_data.source or "MANUAL APPOINTMENT",
                "payment_status": appointment_data.payment_status or "UNPAID",
                "bill_number": bill_ref,
                "clinical_notes": appointment_data.clinical_notes or "",
                "status": str(stat_val),
                "reason": appointment_data.reason or "",
                "notes": appointment_data.notes or "",
                "created_at": now_iso,
                "updated_at": now_iso
            }
            apt_rows.append(new_row)
            all_data[SHEET_APPOINTMENTS] = apt_rows

            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "staff_reception",
                "entity_type": "APPOINTMENT",
                "entity_id": apt_id,
                "action": "CREATE",
                "details": (
                    f"Booked {treatment} for patient {appointment_data.patient_id} "
                    f"with dentist {target_dentist_id} "
                    f"on {appointment_data.date} "
                    f"({appointment_data.start_time}-{appointment_data.end_time})"
                )
            })
            all_data[SHEET_AUDIT] = audit_rows

            self._write_all_sheets(all_data)

        p_info = patients_map.get(appointment_data.patient_id, {})
        dentist_display_name = dentists_map.get(target_dentist_id) or dentists_map.get(appointment_data.dentist_id)
        return AppointmentResponse(
            appointment_id=apt_id,
            patient_id=appointment_data.patient_id,
            patient_name=p_info.get("full_name"),
            patient_phone=p_info.get("phone"),
            dentist_id=target_dentist_id,
            dentist_name=dentist_display_name,
            date=appointment_data.date,
            start_time=appointment_data.start_time,
            end_time=appointment_data.end_time,
            treatment_name=treatment,
            source=appointment_data.source or "MANUAL APPOINTMENT",
            payment_status=appointment_data.payment_status or "UNPAID",
            bill_number=bill_ref,
            clinical_notes=appointment_data.clinical_notes,
            status=AppointmentStatus(stat_val),
            reason=appointment_data.reason,
            notes=appointment_data.notes,
            booking_time=booking_timestamp,
            created_at=now_iso,
            updated_at=now_iso
        )

    def update_appointment_status(
        self,
        appointment_id: str,
        new_status: AppointmentStatus,
        notes: Optional[str] = None
    ) -> Optional[AppointmentResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            apt_rows = all_data.get(SHEET_APPOINTMENTS, [])
            target = None
            for r in apt_rows:
                if r.get("appointment_id", "").lower() == appointment_id.lower():
                    target = r
                    break
            if not target:
                return None

            status_val = new_status.value if hasattr(new_status, "value") else str(new_status)
            now_iso = datetime.now().isoformat()
            target["status"] = status_val
            target["updated_at"] = now_iso
            if notes:
                existing_notes = target.get("notes", "")
                target["notes"] = f"{existing_notes} | {notes}".strip(" | ")

            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "staff_reception",
                "entity_type": "APPOINTMENT",
                "entity_id": appointment_id,
                "action": f"STATUS_{status_val.upper().replace('-', '_')}",
                "details": f"Status updated to {status_val}"
            })
            all_data[SHEET_AUDIT] = audit_rows

            self._write_all_sheets(all_data)

            # Build response inline from in-memory data — avoids re-acquiring the lock
            return self._get_appointment_unlocked(all_data, appointment_id)

    def reschedule_appointment(
        self,
        appointment_id: str,
        new_date: str,
        new_start_time: str,
        new_end_time: str,
        new_dentist_id: Optional[str] = None,
        notes: Optional[str] = None
    ) -> Optional[AppointmentResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            apt_rows = all_data.get(SHEET_APPOINTMENTS, [])
            target = None
            for r in apt_rows:
                if r.get("appointment_id", "").lower() == appointment_id.lower():
                    target = r
                    break
            if not target:
                return None

            effective_dentist_id = new_dentist_id or target.get("dentist_id")

            # Conflict check for the new slot — using in-memory data, no extra lock
            for r in apt_rows:
                if r.get("appointment_id", "").lower() == appointment_id.lower():
                    continue  # skip self
                if (
                    r.get("dentist_id", "").lower() == effective_dentist_id.lower()
                    and r.get("date") == new_date
                    and r.get("status", "").lower() in ["confirmed", "pending"]
                ):
                    existing_start = r.get("start_time", "")
                    existing_end = r.get("end_time", "")
                    if not (new_end_time <= existing_start or new_start_time >= existing_end):
                        raise SlotConflictError(
                            f"Cannot reschedule: Dentist is already booked on {new_date} "
                            f"from {existing_start} to {existing_end}."
                        )

            now_iso = datetime.now().isoformat()
            old_info = (
                f"Old: {target['date']} "
                f"{target['start_time']}-{target['end_time']} "
                f"({target['dentist_id']})"
            )

            target["date"] = new_date
            target["start_time"] = new_start_time
            target["end_time"] = new_end_time
            target["dentist_id"] = effective_dentist_id
            target["status"] = AppointmentStatus.RESCHEDULED.value
            target["updated_at"] = now_iso

            reschedule_note = f"Rescheduled from {old_info} on {now_iso}"
            if notes:
                reschedule_note += f" - Reason: {notes}"
            existing_apt_notes = target.get("notes", "")
            target["notes"] = f"{existing_apt_notes} | {reschedule_note}".strip(" | ")

            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "staff_reception",
                "entity_type": "APPOINTMENT",
                "entity_id": appointment_id,
                "action": "RESCHEDULE",
                "details": (
                    f"{old_info} -> New: {new_date} "
                    f"{new_start_time}-{new_end_time} ({effective_dentist_id})"
                )
            })
            all_data[SHEET_AUDIT] = audit_rows

            self._write_all_sheets(all_data)

            # Build response inline from in-memory data — avoids re-acquiring the lock
            return self._get_appointment_unlocked(all_data, appointment_id)

    # =========================================================================
    # AUDIT LOG
    # =========================================================================

    def log_audit_event(self, entry: AuditLogEntry) -> None:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": entry.timestamp or datetime.now().isoformat(),
                "staff_id": entry.staff_id,
                "entity_type": entry.entity_type,
                "entity_id": entry.entity_id,
                "action": entry.action,
                "details": entry.details
            })
            all_data[SHEET_AUDIT] = audit_rows
            self._write_all_sheets(all_data)

    # =========================================================================
    # TREATMENTS CATALOG
    # =========================================================================

    def list_treatments(self) -> List[Treatment]:
        return [Treatment(**t) for t in DEFAULT_TREATMENTS]

    # =========================================================================
    # BILLING & PAYMENT REMINDERS
    # =========================================================================

    def update_payment_status(
        self,
        appointment_id: str,
        payment_status: str,
        bill_number: Optional[str] = None
    ) -> Optional[AppointmentResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            apt_rows = all_data.get(SHEET_APPOINTMENTS, [])
            target = None
            for r in apt_rows:
                if r.get("appointment_id", "").lower() == appointment_id.lower():
                    target = r
                    break
            if not target:
                return None

            now_iso = datetime.now().isoformat()
            target["payment_status"] = payment_status
            if bill_number:
                target["bill_number"] = bill_number
            target["updated_at"] = now_iso

            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "staff_reception",
                "entity_type": "APPOINTMENT",
                "entity_id": appointment_id,
                "action": "PAYMENT_UPDATE",
                "details": f"Updated payment status to {payment_status} (Bill: {target.get('bill_number', '')})"
            })
            all_data[SHEET_AUDIT] = audit_rows
            self._write_all_sheets(all_data)

            return self._get_appointment_unlocked(all_data, appointment_id)

    def send_payment_reminder(self, appointment_id: str) -> Optional[PaymentReminderResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            apt = self._get_appointment_unlocked(all_data, appointment_id)
            if not apt:
                return None

            now_iso = datetime.now().isoformat()
            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "staff_reception",
                "entity_type": "APPOINTMENT",
                "entity_id": appointment_id,
                "action": "SEND_REMINDER",
                "details": f"Payment reminder triggered for patient {apt.patient_id} ({apt.bill_number or 'Unbilled'})"
            })
            all_data[SHEET_AUDIT] = audit_rows
            self._write_all_sheets(all_data)

        patient_name = apt.patient_name or "Patient"
        bill_ref = apt.bill_number or "Pending Bill"
        msg = f"Dear {patient_name}, reminder from SmileCare Dental: your payment for {apt.treatment_name or 'procedure'} ({bill_ref}) is pending. Please contact the front desk."

        return PaymentReminderResponse(
            appointment_id=apt.appointment_id,
            patient_id=apt.patient_id,
            patient_name=apt.patient_name,
            patient_phone=apt.patient_phone,
            bill_number=apt.bill_number,
            amount_or_status=apt.payment_status or "UNPAID",
            reminder_sent_at=now_iso,
            channel="SMS/WhatsApp Simulator",
            message=msg
        )

    # =========================================================================
    # MEDICAL CHECKUP & ODONTOGRAM
    # =========================================================================

    def _deserialize_checkup_row(self, r: Dict[str, Any]) -> MedicalCheckupResponse:
        findings_raw = r.get("teeth_findings_json", "")
        teeth_findings = []
        if findings_raw:
            try:
                parsed = json.loads(findings_raw)
                teeth_findings = [ToothFinding(**item) for item in parsed]
            except Exception:
                teeth_findings = []

        conditions_raw = r.get("medical_conditions", "")
        medical_conditions = []
        if conditions_raw:
            try:
                medical_conditions = json.loads(conditions_raw)
            except Exception:
                medical_conditions = [c.strip() for c in conditions_raw.split(",") if c.strip()]

        canker_bool = str(r.get("canker_sores", "FALSE")).upper() in ["TRUE", "1", "YES"]
        anomalous_bool = str(r.get("anomalous_teeth", "FALSE")).upper() in ["TRUE", "1", "YES"]

        return MedicalCheckupResponse(
            checkup_id=r.get("checkup_id", ""),
            patient_id=r.get("patient_id", ""),
            appointment_id=r.get("appointment_id") or None,
            dentist_id=r.get("dentist_id") or None,
            dentist_name=r.get("dentist_name") or None,
            blood_pressure=r.get("blood_pressure") or None,
            medical_conditions=medical_conditions,
            allergies=r.get("allergies") or None,
            oral_hygiene_habits=r.get("oral_hygiene_habits") or None,
            teeth_findings=teeth_findings,
            canker_sores=canker_bool,
            canker_sores_notes=r.get("canker_sores_notes") or None,
            anomalous_teeth=anomalous_bool,
            anomalous_teeth_notes=r.get("anomalous_teeth_notes") or None,
            other_oral_notes=r.get("other_oral_notes") or None,
            consent_status=r.get("consent_status", "approved"),
            refusal_reason=r.get("refusal_reason") or None,
            document_url_or_ref=r.get("document_url_or_ref") or None,
            status=r.get("status", "completed"),
            created_at=r.get("created_at", ""),
            updated_at=r.get("updated_at", "")
        )

    def save_medical_checkup(self, checkup_data: MedicalCheckupCreate) -> MedicalCheckupResponse:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            checkup_rows = all_data.get(SHEET_CHECKUPS, [])
            now_iso = datetime.now().isoformat()

            # Serialize teeth findings and medical conditions
            teeth_json = json.dumps([t.model_dump() if hasattr(t, "model_dump") else t.dict() for t in checkup_data.teeth_findings])
            conditions_json = json.dumps(checkup_data.medical_conditions)

            # Check if updating existing checkup for this appointment
            target = None
            if checkup_data.appointment_id:
                for r in checkup_rows:
                    if r.get("appointment_id", "").lower() == checkup_data.appointment_id.lower():
                        target = r
                        break

            if target:
                checkup_id = target["checkup_id"]
                target["dentist_id"] = checkup_data.dentist_id or target.get("dentist_id", "")
                target["blood_pressure"] = checkup_data.blood_pressure or ""
                target["medical_conditions"] = conditions_json
                target["allergies"] = checkup_data.allergies or ""
                target["oral_hygiene_habits"] = checkup_data.oral_hygiene_habits or ""
                target["teeth_findings_json"] = teeth_json
                target["canker_sores"] = "TRUE" if checkup_data.canker_sores else "FALSE"
                target["canker_sores_notes"] = checkup_data.canker_sores_notes or ""
                target["anomalous_teeth"] = "TRUE" if checkup_data.anomalous_teeth else "FALSE"
                target["anomalous_teeth_notes"] = checkup_data.anomalous_teeth_notes or ""
                target["other_oral_notes"] = checkup_data.other_oral_notes or ""
                target["consent_status"] = checkup_data.consent_status
                target["refusal_reason"] = checkup_data.refusal_reason or ""
                target["status"] = checkup_data.status
                target["updated_at"] = now_iso
            else:
                seq = self._next_sequence(checkup_rows, "checkup_id")
                checkup_id = f"CHK-{seq:06d}"
                new_row = {
                    "checkup_id": checkup_id,
                    "patient_id": checkup_data.patient_id,
                    "appointment_id": checkup_data.appointment_id or "",
                    "dentist_id": checkup_data.dentist_id or "",
                    "blood_pressure": checkup_data.blood_pressure or "",
                    "medical_conditions": conditions_json,
                    "allergies": checkup_data.allergies or "",
                    "oral_hygiene_habits": checkup_data.oral_hygiene_habits or "",
                    "teeth_findings_json": teeth_json,
                    "canker_sores": "TRUE" if checkup_data.canker_sores else "FALSE",
                    "canker_sores_notes": checkup_data.canker_sores_notes or "",
                    "anomalous_teeth": "TRUE" if checkup_data.anomalous_teeth else "FALSE",
                    "anomalous_teeth_notes": checkup_data.anomalous_teeth_notes or "",
                    "other_oral_notes": checkup_data.other_oral_notes or "",
                    "consent_status": checkup_data.consent_status,
                    "refusal_reason": checkup_data.refusal_reason or "",
                    "status": checkup_data.status,
                    "created_at": now_iso,
                    "updated_at": now_iso
                }
                checkup_rows.append(new_row)
                target = new_row

            all_data[SHEET_CHECKUPS] = checkup_rows

            # Also synchronize clinical summary banner to linked appointment
            if checkup_data.appointment_id:
                for apt_r in all_data.get(SHEET_APPOINTMENTS, []):
                    if apt_r.get("appointment_id", "").lower() == checkup_data.appointment_id.lower():
                        summary_note = checkup_data.canker_sores_notes or checkup_data.other_oral_notes or "Checkup completed"
                        apt_r["clinical_notes"] = summary_note
                        if checkup_data.status == "completed":
                            apt_r["status"] = "finished"
                        apt_r["updated_at"] = now_iso
                        break

            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "staff_dentist",
                "entity_type": "CHECKUP",
                "entity_id": checkup_id,
                "action": "SAVE",
                "details": f"Saved medical checkup and odontogram for patient {checkup_data.patient_id} (findings: {len(checkup_data.teeth_findings)} teeth)"
            })
            all_data[SHEET_AUDIT] = audit_rows

            self._write_all_sheets(all_data)

            return self._deserialize_checkup_row(target)

    def get_medical_checkup(self, checkup_id: str) -> Optional[MedicalCheckupResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            for r in all_data.get(SHEET_CHECKUPS, []):
                if r.get("checkup_id", "").lower() == checkup_id.lower():
                    return self._deserialize_checkup_row(r)
            return None

    def get_checkup_by_appointment(self, appointment_id: str) -> Optional[MedicalCheckupResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            for r in all_data.get(SHEET_CHECKUPS, []):
                if r.get("appointment_id", "").lower() == appointment_id.lower():
                    return self._deserialize_checkup_row(r)
            return None

    def list_checkups_for_patient(self, patient_id: str) -> List[MedicalCheckupResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            results = []
            for r in all_data.get(SHEET_CHECKUPS, []):
                if r.get("patient_id", "").lower() == patient_id.lower():
                    results.append(self._deserialize_checkup_row(r))
            return sorted(results, key=lambda c: c.created_at, reverse=True)

    # =========================================================================
    # PATIENT REQUESTS (SIMULATOR / WHATSAPP)
    # =========================================================================

    def _deserialize_patient_request_row(self, row: Dict[str, Any]) -> PatientRequestResponse:
        return PatientRequestResponse(
            request_id=str(row.get("request_id", "")),
            patient_name=str(row.get("patient_name", "")),
            patient_phone=str(row.get("patient_phone", "")),
            patient_age=str(row["patient_age"]) if row.get("patient_age") else None,
            patient_id=str(row["patient_id"]) if row.get("patient_id") else None,
            dentist_id=str(row.get("dentist_id", "")),
            preferred_date=str(row.get("preferred_date", "")),
            preferred_start_time=str(row.get("preferred_start_time", "")),
            preferred_end_time=str(row.get("preferred_end_time", "")),
            reason=str(row["reason"]) if row.get("reason") else None,
            source=str(row.get("source", "simulator")),
            status=PatientRequestStatus(str(row.get("status", "pending"))),
            review_notes=str(row["review_notes"]) if row.get("review_notes") else None,
            appointment_id=str(row["appointment_id"]) if row.get("appointment_id") else None,
            booking_time=str(row.get("booking_time") or row.get("created_at", "")),
            created_at=str(row.get("created_at", datetime.now().isoformat())),
            updated_at=str(row.get("updated_at", datetime.now().isoformat()))
        )

    def create_patient_request(self, request_data: PatientRequestCreate) -> PatientRequestResponse:
        now_iso = datetime.now().isoformat()
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            req_rows = all_data.get(SHEET_PATIENT_REQUESTS, [])
            seq = self._next_sequence(req_rows, "request_id")
            request_id = f"REQ-{seq:06d}"

            new_row = {
                "request_id": request_id,
                "patient_name": request_data.patient_name,
                "patient_phone": request_data.patient_phone,
                "patient_age": request_data.patient_age or "",
                "patient_id": request_data.patient_id or "",
                "dentist_id": request_data.dentist_id,
                "preferred_date": request_data.preferred_date,
                "preferred_start_time": request_data.preferred_start_time,
                "preferred_end_time": request_data.preferred_end_time,
                "booking_time": now_iso,
                "reason": request_data.reason or "",
                "source": request_data.source or "simulator",
                "status": PatientRequestStatus.PENDING.value,
                "review_notes": "",
                "appointment_id": "",
                "created_at": now_iso,
                "updated_at": now_iso
            }
            req_rows.append(new_row)
            all_data[SHEET_PATIENT_REQUESTS] = req_rows

            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "simulator_bot",
                "entity_type": "PATIENT_REQUEST",
                "entity_id": request_id,
                "action": "CREATE",
                "details": f"Patient request submitted from {request_data.source} for {request_data.patient_name}"
            })
            all_data[SHEET_AUDIT] = audit_rows

            self._write_all_sheets(all_data)
            return self._deserialize_patient_request_row(new_row)

    def get_patient_request(self, request_id: str) -> Optional[PatientRequestResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            for r in all_data.get(SHEET_PATIENT_REQUESTS, []):
                if r.get("request_id", "").lower() == request_id.lower():
                    return self._deserialize_patient_request_row(r)
            return None

    def list_patient_requests(self, status: Optional[str] = None) -> List[PatientRequestResponse]:
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            results = []
            for r in all_data.get(SHEET_PATIENT_REQUESTS, []):
                if status:
                    if r.get("status", "").lower() != status.lower():
                        continue
                results.append(self._deserialize_patient_request_row(r))
            return sorted(results, key=lambda x: x.created_at, reverse=True)

    def update_patient_request_status(
        self, request_id: str, status: str, review_notes: Optional[str] = None, appointment_id: Optional[str] = None
    ) -> Optional[PatientRequestResponse]:
        now_iso = datetime.now().isoformat()
        with get_workbook_lock(self.lock_path):
            all_data = self._read_all_sheets()
            req_rows = all_data.get(SHEET_PATIENT_REQUESTS, [])
            target = None
            for r in req_rows:
                if r.get("request_id", "").lower() == request_id.lower():
                    r["status"] = status
                    if review_notes is not None:
                        r["review_notes"] = review_notes
                    if appointment_id is not None:
                        r["appointment_id"] = appointment_id
                    r["updated_at"] = now_iso
                    target = r
                    break

            if not target:
                return None

            all_data[SHEET_PATIENT_REQUESTS] = req_rows

            audit_rows = all_data.get(SHEET_AUDIT, [])
            audit_rows.append({
                "log_id": f"LOG-{self._next_sequence(audit_rows, 'log_id'):06d}",
                "timestamp": now_iso,
                "staff_id": "staff_desk",
                "entity_type": "PATIENT_REQUEST",
                "entity_id": request_id,
                "action": "UPDATE_STATUS",
                "details": f"Updated request status to {status}. Notes: {review_notes or 'None'}"
            })
            all_data[SHEET_AUDIT] = audit_rows

            self._write_all_sheets(all_data)
            return self._deserialize_patient_request_row(target)

    # ── Fallback Implementations for New Abstract Methods ───
    def delete_patient(self, patient_id: str) -> bool:
        return False

    def update_appointment(self, appointment_id: str, updates: Dict[str, Any]) -> Optional[AppointmentResponse]:
        return self.get_appointment(appointment_id)

    def get_treatment(self, treatment_id: str) -> Optional[Treatment]:
        if treatment_id in self._treatments:
            return self._treatments[treatment_id]
        for t in self.list_treatments():
            if t.treatment_id == treatment_id:
                return t
        return None

    def create_treatment(self, treatment_data: TreatmentCreate) -> Treatment:
        tid = f"TRT-{len(self._treatments) + 10:06d}"
        t = Treatment(
            treatment_id=tid,
            name=treatment_data.name,
            category=treatment_data.category or "General",
            default_duration_minutes=treatment_data.default_duration_minutes or 30,
            estimated_cost=treatment_data.estimated_cost,
            description=treatment_data.description
        )
        self._treatments[tid] = t
        return t

    def update_treatment(self, treatment_id: str, updates: TreatmentUpdate) -> Optional[Treatment]:
        t = self.get_treatment(treatment_id)
        if not t:
            return None
        t_dict = t.model_dump()
        t_dict.update(updates.model_dump(exclude_unset=True))
        updated = Treatment(**t_dict)
        self._treatments[treatment_id] = updated
        return updated

    def delete_treatment(self, treatment_id: str) -> bool:
        if treatment_id in self._treatments:
            del self._treatments[treatment_id]
        return True

    def list_staff(self, active_only: bool = False) -> List[StaffMember]:
        base_staff = [
            StaffMember(staff_id="STF-000001", username="admin", full_name="Emma Watson", role="Clinic Manager", department="Administration", initials="EW", status="Active", is_active=True),
            StaffMember(staff_id="STF-000002", username="receptionist1", full_name="Jessica Taylor", role="Head Receptionist", department="Front Desk", initials="JT", status="Active", is_active=True),
            StaffMember(staff_id="STF-000003", username="nurse1", full_name="Alex Robinson", role="Senior Dental Assistant", department="Nursing", initials="AR", status="Active", is_active=True),
        ]
        all_s = {s.staff_id: s for s in base_staff}
        all_s.update(self._staff)
        res = list(all_s.values())
        if active_only:
            res = [s for s in res if s.is_active and s.status == "Active"]
        return res

    def get_staff(self, staff_id: str) -> Optional[StaffMember]:
        if staff_id in self._staff:
            return self._staff[staff_id]
        for s in self.list_staff():
            if s.staff_id == staff_id:
                return s
        return None

    def create_staff(self, staff_data: StaffCreate) -> StaffMember:
        sid = f"STF-{len(self._staff) + 10:06d}"
        parts = staff_data.full_name.strip().split()
        initials = "".join([p[0].upper() for p in parts if p])[:2] or "ST"
        s = StaffMember(
            staff_id=sid,
            full_name=staff_data.full_name,
            role=staff_data.role,
            department=staff_data.department,
            phone=staff_data.phone,
            email=staff_data.email,
            initials=initials,
            status="Active",
            is_active=True
        )
        self._staff[sid] = s
        return s

    def update_staff(self, staff_id: str, updates: StaffUpdate) -> Optional[StaffMember]:
        s = self.get_staff(staff_id)
        if not s:
            return None
        d = s.model_dump()
        d.update(updates.model_dump(exclude_unset=True))
        updated = StaffMember(**d)
        self._staff[staff_id] = updated
        return updated

    def delete_staff(self, staff_id: str) -> bool:
        if staff_id in self._staff:
            del self._staff[staff_id]
        return True

    def list_sales(self, date: Optional[str] = None, patient_id: Optional[str] = None, status: Optional[str] = None) -> List[SaleResponse]:
        res = list(self._sales.values())
        if status:
            res = [s for s in res if s.status.lower() == status.lower()]
        return res

    def get_sale(self, sale_id: str) -> Optional[SaleResponse]:
        return self._sales.get(sale_id)

    def create_sale(self, sale_data: SaleCreate) -> SaleResponse:
        sid = f"SAL-{len(self._sales) + 1:06d}"
        s = SaleResponse(
            sale_id=sid,
            patient_name=sale_data.patient_name,
            treatment_name=sale_data.treatment_name,
            amount=sale_data.amount,
            status=sale_data.status or "Pending",
            payment_method=sale_data.payment_method or "Cash",
            notes=sale_data.notes,
            created_at=datetime.now().isoformat()
        )
        self._sales[sid] = s
        return s

    def update_sale_status(self, sale_id: str, status: str) -> Optional[SaleResponse]:
        s = self._sales.get(sale_id)
        if s:
            d = s.model_dump()
            d["status"] = status
            updated = SaleResponse(**d)
            self._sales[sale_id] = updated
            return updated
        return None

    def get_sales_summary(self) -> SaleSummary:
        sales = list(self._sales.values())
        total_revenue = sum(s.amount for s in sales if s.status == "Paid")
        total_pending = sum(s.amount for s in sales if s.status != "Paid")
        return SaleSummary(
            total_sales=len(sales),
            total_paid=total_revenue,
            total_pending=total_pending
        )

    def list_vendors(self) -> List[VendorResponse]:
        return list(self._vendors.values())

    def create_vendor(self, vendor_data: VendorCreate) -> VendorResponse:
        vid = f"VND-{len(self._vendors) + 1:06d}"
        v = VendorResponse(
            vendor_id=vid,
            name=vendor_data.name,
            contact=vendor_data.contact,
            email=vendor_data.email,
            phone=vendor_data.phone,
            address=vendor_data.address,
            created_at=datetime.now().isoformat()
        )
        self._vendors[vid] = v
        return v

    def list_purchases(self, status: Optional[str] = None) -> List[PurchaseResponse]:
        res = list(self._purchases.values())
        if status:
            res = [p for p in res if p.status.lower() == status.lower()]
        return res

    def get_purchase(self, purchase_id: str) -> Optional[PurchaseResponse]:
        return self._purchases.get(purchase_id)

    def create_purchase(self, purchase_data: PurchaseCreate) -> PurchaseResponse:
        pid = f"PO-{len(self._purchases) + 1:06d}"
        p = PurchaseResponse(
            purchase_id=pid,
            vendor_name=purchase_data.vendor_name,
            items=purchase_data.items,
            amount=purchase_data.amount,
            status=purchase_data.status or "Ordered",
            created_at=datetime.now().isoformat()
        )
        self._purchases[pid] = p
        return p

    def update_purchase_status(self, purchase_id: str, status: str, received_date: Optional[str] = None) -> Optional[PurchaseResponse]:
        p = self._purchases.get(purchase_id)
        if p:
            d = p.model_dump()
            d["status"] = status
            if received_date:
                d["received_date"] = received_date
            updated = PurchaseResponse(**d)
            self._purchases[purchase_id] = updated
            return updated
        return None

    def list_inventory(self, category: Optional[str] = None, low_stock_only: bool = False) -> List[InventoryResponse]:
        res = list(self._inventory.values())
        if category:
            res = [i for i in res if i.category.lower() == category.lower()]
        if low_stock_only:
            res = [i for i in res if i.quantity <= i.min_stock]
        return res

    def get_inventory_item(self, item_id: str) -> Optional[InventoryResponse]:
        return self._inventory.get(item_id)

    def create_inventory_item(self, item_data: InventoryCreate) -> InventoryResponse:
        iid = f"INV-{len(self._inventory) + 1:06d}"
        item = InventoryResponse(
            item_id=iid,
            name=item_data.name,
            category=item_data.category or "Consumables",
            quantity=item_data.quantity or 0,
            min_stock=item_data.min_stock or 0,
            unit=item_data.unit or "pcs",
            unit_price=item_data.unit_price or 0.0,
            supplier=item_data.supplier,
            created_at=datetime.now().isoformat()
        )
        self._inventory[iid] = item
        return item

    def update_inventory_item(self, item_id: str, updates: InventoryUpdate) -> Optional[InventoryResponse]:
        item = self._inventory.get(item_id)
        if item:
            d = item.model_dump()
            d.update(updates.model_dump(exclude_unset=True))
            updated = InventoryResponse(**d)
            self._inventory[item_id] = updated
            return updated
        return None

    def delete_inventory_item(self, item_id: str) -> bool:
        if item_id in self._inventory:
            del self._inventory[item_id]
        return True

    def list_payment_methods(self) -> List[PaymentMethodResponse]:
        return list(self._payment_methods.values())

    def update_payment_method(self, method_id: str, enabled: Optional[bool] = None, processing_fee: Optional[str] = None) -> Optional[PaymentMethodResponse]:
        pm = self._payment_methods.get(method_id)
        if pm:
            d = pm.model_dump()
            if enabled is not None:
                d["is_active"] = enabled
            if processing_fee is not None:
                d["processing_fee"] = processing_fee
            updated = PaymentMethodResponse(**d)
            self._payment_methods[method_id] = updated
            return updated
        return None

    def _ensure_runtime_defaults(self) -> None:
        if not self._vendors:
            for name in ["DentSupply Co.", "Medix Pharma", "BioTech Dental", "Global Dental Direct"]:
                self.create_vendor(VendorCreate(name=name))
        if not self._peripherals:
            defaults = [
                ("Dental Chair #1", "Chair", "Room 1", "Good", "DC-2021-001", "2024-02-15"),
                ("Dental Chair #2", "Chair", "Room 2", "Good", "DC-2021-002", "2024-02-15"),
                ("Dental X-Ray Machine", "Imaging", "Room 1", "Good", "XR-2020-007", "2024-03-20"),
                ("Digital Panoramic X-Ray", "Imaging", "X-Ray", "Service", "PX-2019-003", "2023-11-10"),
                ("Autoclave Sterilizer", "Sterilization", "Lab", "Good", "AC-2022-001", "2024-01-08"),
                ("Intraoral Camera", "Imaging", "Room 2", "Good", "IC-2023-005", "2024-04-05"),
                ("Dental Compressor", "Equipment", "Utility", "Good", "CP-2021-002", "2024-01-22"),
                ("Patient Monitor #1", "Monitor", "Room 1", "Good", "PM-2022-001", "2024-03-15"),
                ("Reception Computer", "IT", "Front", "Good", "PC-2023-001", "2024-04-20"),
                ("Billing Printer", "IT", "Front", "Needs Check", "PR-2020-004", "2023-08-01"),
            ]
            for name, category, location, condition, serial_no, last_service in defaults:
                self.create_peripheral(PeripheralCreate(
                    name=name,
                    category=category,
                    location=location,
                    condition=condition,
                    serial_no=serial_no,
                    last_service=last_service,
                ))

    def list_peripherals(self) -> List[PeripheralResponse]:
        return list(self._peripherals.values())

    def get_peripheral(self, peripheral_id: str) -> Optional[PeripheralResponse]:
        return self._peripherals.get(peripheral_id)

    def create_peripheral(self, item_data: PeripheralCreate) -> PeripheralResponse:
        pid = f"PER-{len(self._peripherals) + 1:06d}"
        item = PeripheralResponse(
            peripheral_id=pid,
            name=item_data.name,
            category=item_data.category or "Equipment",
            location=item_data.location or "",
            condition=item_data.condition or "Good",
            serial_no=item_data.serial_no,
            last_service=item_data.last_service,
            created_at=datetime.now().isoformat(),
        )
        self._peripherals[pid] = item
        return item

    def update_peripheral(self, peripheral_id: str, updates: PeripheralUpdate) -> Optional[PeripheralResponse]:
        item = self._peripherals.get(peripheral_id)
        if not item:
            return None
        d = item.model_dump()
        d.update(updates.model_dump(exclude_unset=True))
        updated = PeripheralResponse(**d)
        self._peripherals[peripheral_id] = updated
        return updated

    def delete_peripheral(self, peripheral_id: str) -> bool:
        if peripheral_id in self._peripherals:
            del self._peripherals[peripheral_id]
            return True
        return False




